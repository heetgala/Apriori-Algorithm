from openpyxl import *
from tkinter import *
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('data.xlsx') 
  
# create the sheet object 
sheet = wb.active 
  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
        # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "0"
    sheet.cell(row=1, column=2).value = "1"
    sheet.cell(row=1, column=3).value = "2"
    sheet.cell(row=1, column=4).value = "3"
   
  
  

  
  
# Function to set focus 
def focus1(event): 
    # set focus on the sem_field box 
    product1_field.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the form_no_field box 
    product2_field.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the contact_no_field box 
    product3_field.focus_set() 
  
  
# Function to set focus 

  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    product0_field.delete(0, END) 
    product1_field.delete(0, END) 
    product2_field.delete(0, END) 
    product3_field.delete(0, END) 
      
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (product0_field.get() == "" and
        product1_field.get() == "" and
        product2_field.get() == "" and
        product3_field.get() == "" ):
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = product0_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = product1_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = product2_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = product3_field.get() 
         
        # save the file 
        wb.save('data.xlsx') 
  
        # set focus on the name_field box 
        product0_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("Product Details") 
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
    excel() 
  
    # create a Form label 
    heading = Label(root, text="Mhm Store", bg="light green") 
  
    # create a Name label 
    product0 = Label(root, text="Product", bg="light green") 
  
    # create a Course label 
    product1 = Label(root, text="Product", bg="light green") 
  
    # create a Semester label 
    product2 = Label(root, text="Product", bg="light green") 
  
    # create a Form No. lable 
    product3 = Label(root, text="Product", bg="light green") 
  
    
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    product0.grid(row=1, column=0) 
    product1.grid(row=2, column=0) 
    product2.grid(row=3, column=0) 
    product3.grid(row=4, column=0) 
   
  
    # create a text entry box 
    # for typing the information 
    product0_field = Entry(root) 
    product1_field = Entry(root) 
    product2_field = Entry(root) 
    product3_field = Entry(root) 
    
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    product0_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    product1_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    product2_field.bind("<Return>", focus3) 
  
    
  
      # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    product0_field.grid(row=1, column=1, ipadx="100") 
    product1_field.grid(row=2, column=1, ipadx="100") 
    product2_field.grid(row=3, column=1, ipadx="100") 
    product3_field.grid(row=4, column=1, ipadx="100") 
       # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=8, column=1) 
  
    # start the GUI 
    root.mainloop() 

